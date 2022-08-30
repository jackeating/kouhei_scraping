from selenium import webdriver
import openpyxl
from datetime import datetime

def main():
    driver = webdriver.Chrome(executable_path="./chromedriver")
    prefecture_dict = {
        "北海道": "1",
        "青森県": "2",
        "岩手県": "3",
        "宮城県": "4",
        "秋田県": "5",
        "山形県": "6",
        "福島県": "7",
        "東京都": "13",
        "神奈川県": "14",
        "埼玉県": "11",
        "千葉県": "12",
        "茨城県": "8",
        "栃木県": "9",
        "群馬県": "10",
        "新潟県": "15",
        "長野県": "20",
        "富山県": "16",
        "山梨県": "19",
        "石川県": "17",
        "福井県": "18",
        "愛知県": "23",
        "岐阜県": "21",
        "静岡県": "22",
        "三重県": "24",
        "大阪府": "27",
        "兵庫県": "28",
        "京都府": "26",
        "滋賀県": "25",
        "奈良県": "29",
        "和歌山県": "30",
        "鳥取県": "31",
        "島根県": "32",
        "岡山県": "33",
        "広島県": "34",
        "山口県": "35",
        "徳島県": "36",
        "香川県": "37",
        "愛媛県": "38",
        "高知県": "39",
        "福岡県": "40",
        "佐賀県": "41",
        "長崎県": "42",
        "熊本県": "43",
        "大分県": "44",
        "宮崎県": "45",
        "鹿児島県": "46",
        "沖縄県": "47"
    }
    result_dict = {}
    for prefecture, num in prefecture_dict.items():
        # 都道府県別市区町村の全URLを格納する変数
        url_each_municipality = []
        # 都道府県毎のdict作成
        result_dict[prefecture] = {}
        # 都道府県別に市区町村を取得
        driver.get("http://www.hanakara.jp/search_hospital/select_address.html?area1=" + num)
        elems = driver.find_elements_by_xpath("//a[@href]")
        for elem in elems:
            try:
                all_href = elem.get_attribute("href")
            except:
                continue
            if "disp_list" not in all_href:
                    continue
            url_each_municipality.append(all_href)
        
        for all_href in url_each_municipality:
            driver.get(all_href)
            municipalities_list = []
            municipalitie_name = driver.find_element_by_xpath("//*[@id='js-search-section-title-name']").text
            municipalitie_name = municipalitie_name[:-1]
            result_dict[prefecture][municipalitie_name] = []
            # 区町村ごとの病院を取得
            municipalities = driver.find_elements_by_xpath("//a[@href]")
            for municipalitie in municipalities:
                try:
                    municipalitie_data = municipalitie.get_attribute("href")
                except:
                    continue
                if "disp_detail" not in municipalitie_data:
                    continue
                municipalities_list.append(municipalitie_data)
            for municipalitie_data in municipalities_list:
                driver.get(municipalitie_data)
                search_detail_name = driver.find_element_by_xpath("//*[@id='search_detail_name']").text
                search_detail_address = driver.find_element_by_xpath("//*[@id='search_detail_address']").text
                # search_detail_tel = driver.find_element_by_xpath("//*[@id='search_detail_tel']").text
                # search_detail_kamoku = driver.find_element_by_xpath("//*[@id='search_detail_kamoku']").text
                # search_detail_hour = driver.find_element_by_xpath("//*[@id='search_detail_hour']").text
                # search_detail_closed = driver.find_element_by_xpath("//*[@id='search_detail_closed']").text
                # search_detail_url = driver.find_element_by_xpath("//*[@id='search_detail_url']").text
                result_dict[prefecture][municipalitie_name].append(
                    {
                        "病院名": search_detail_name,
                        "住所": search_detail_address,
                        # "TEL": search_detail_tel,
                        # "診療科目": search_detail_kamoku,
                        # "診療時間": search_detail_hour,
                        # "休診日": search_detail_closed,
                        # "URL": search_detail_url,
                    }
                )
            break
        break
    wb=openpyxl.Workbook()
    ws=wb.active
    # シート名の設定
    ws.title="test"
    # シートの読み込み
    sheet = wb['test']
    # 値の代入
    sheet.cell(column=1,row=1).value = "都道府県"
    sheet.cell(column=2,row=1).value = "市区町村"
    sheet.cell(column=3,row=1).value = "病院名"
    sheet.cell(column=4,row=1).value = "住所"

    row_num = 1
    for prefecture, municipalities_dict in result_dict.items():
        for municipalities, hospital_list in municipalities_dict.items():
            for hospital_data in hospital_list:
                row_num += 1
                sheet.cell(column=1,row=row_num).value = prefecture
                sheet.cell(column=2,row=row_num).value = municipalities
                sheet.cell(column=3,row=row_num).value = hospital_data.get("病院名")
                sheet.cell(column=4,row=row_num).value = hospital_data.get("住所")
    today = datetime.today()
    wb.save('./{}年{}月_data.xlsx'.format(str(today.year), str(today.month)))
if __name__ == '__main__':
    main()